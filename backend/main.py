from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Depends, Body
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List
import openpyxl
import os
from pydantic import BaseModel

app = FastAPI()

# Allow CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = "potluck_data.xlsx"
ADMIN_PASSWORD = "admin123"  # Change this in production!

CATEGORIES = [
    "Starters",
    "Salads",
    "Veg curry",
    "Non-veg Curry",
    "Chapatis/Naan/Roti etc",
    "Rice Items",
    "Sweets",
    "Drinks"
]

class Entry(BaseModel):
    name: str
    category: str
    dish: str
    quantity: int

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Category", "Dish", "Quantity"])
        wb.save(EXCEL_FILE)

@app.on_event("startup")
def startup_event():
    init_excel()

@app.post("/submit")
def submit_entry(entry: Entry):
    if entry.category not in CATEGORIES:
        raise HTTPException(status_code=400, detail="Invalid category")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([entry.name, entry.category, entry.dish, entry.quantity])
    wb.save(EXCEL_FILE)
    return {"message": "Entry saved"}

@app.get("/entries", response_model=List[Entry])
def get_entries():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        entries.append(Entry(name=row[0], category=row[1], dish=row[2], quantity=row[3]))
    return entries

@app.post("/admin/login")
def admin_login(password: str = Form(...)):
    if password == ADMIN_PASSWORD:
        return {"success": True}
    else:
        raise HTTPException(status_code=401, detail="Invalid password")

@app.get("/admin/download")
def download_excel(password: str):
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid password")
    return FileResponse(EXCEL_FILE, filename=EXCEL_FILE)

# Admin edit and delete endpoints
@app.post("/admin/edit")
def admin_edit_entry(password: str = Body(...), index: int = Body(...), name: str = Body(...), category: str = Body(...), dish: str = Body(...), quantity: int = Body(...)):
    if password != ADMIN_PASSWORD:
        return JSONResponse(status_code=401, content={"error": "Unauthorized"})
    entries = get_entries()
    if index < 0 or index >= len(entries):
        return JSONResponse(status_code=400, content={"error": "Invalid index"})
    entries[index] = Entry(name=name, category=category, dish=dish, quantity=quantity)
    # Save all entries to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Category", "Dish", "Quantity"])
    for e in entries:
        ws.append([e.name, e.category, e.dish, e.quantity])
    wb.save(EXCEL_FILE)
    return {"success": True}

@app.post("/admin/delete")
def admin_delete_entry(password: str = Body(...), index: int = Body(...)):
    if password != ADMIN_PASSWORD:
        return JSONResponse(status_code=401, content={"error": "Unauthorized"})
    entries = get_entries()
    if index < 0 or index >= len(entries):
        return JSONResponse(status_code=400, content={"error": "Invalid index"})
    entries.pop(index)
    # Save all entries to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Category", "Dish", "Quantity"])
    for e in entries:
        ws.append([e.name, e.category, e.dish, e.quantity])
    wb.save(EXCEL_FILE)
    return {"success": True}
